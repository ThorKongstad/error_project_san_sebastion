import argparse
import sys
import pathlib
from copy import copy
import traceback

sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))
from error_project_san_sebastion import build_pd
#from . import build_pd
from error_project_san_sebastion.reaction_functions import Functional, get_needed_structures
#from .reaction_functions import Functional, get_needed_structures
from error_project_san_sebastion.reactions import all_gaseous_reactions, all_formation_reactions, all_gaseous_reactions_named, all_formation_reactions_named
#from .reactions import all_gaseous_reactions, all_formation_reactions, all_gaseous_reactions_named, all_formation_reactions_named
from error_project_san_sebastion.error_decomposition import simple_decomposition, lstsq_decomposition
#from .error_decomposition import simple_decomposition, lstsq_decomposition
from error_project_san_sebastion.manual_functional_groups_revised import molecule_functional_dict
#from .manual_functional_groups_revised import molecule_functional_dict

import pandas as pd
import openpyxl as xl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles.borders import Border, Side, BORDER_THIN


def missing_structures(functional: Functional, needed_strcutures: dict):
    missing_molecules = {key for key in needed_strcutures['molecule'] if key not in functional.molecule.keys()}
    missing_adsorbates = {key for key in needed_strcutures['adsorbate'] if key not in functional.adsorbate.keys()}
    missing_slab = {key for key in needed_strcutures['slab'] if key not in functional.slab.keys()}

    if any(len(mis)>0 for mis in (missing_adsorbates, missing_slab, missing_molecules)):
        print(f'{functional.name} is missing')
        if len(missing_molecules) > 0:
            print('    molecules:')
            print(missing_molecules)
        if len(missing_adsorbates) > 0:
            print('    adsorbates:')
            print(missing_adsorbates)
        if len(missing_slab) > 0:
            print('    slab:')
            print(missing_slab)
        print()


def sp(x):
    print(x)
    return x


def main(molecule_database_dir: str, solid_database_dir: str, verbose: bool = False):
    pd_molecule_dat = build_pd(molecule_database_dir)
    pd_solid_dat = build_pd(solid_database_dir)

    pd_molecule_dat['enthalpy'] = pd_molecule_dat['energy'] + pd_molecule_dat['zpe']
    pd_solid_dat['enthalpy'] = pd_solid_dat['energy'] + pd_solid_dat['zpe']

    functional_set = {xc for _, row in pd_molecule_dat.iterrows() if not pd.isna((xc := row.get('xc')))}

    need_structures = get_needed_structures(all_gaseous_reactions + all_formation_reactions)

    prefered_functional_order = {'pbe': 1, 'pw91': 2, 'rpbe': 3, 'beef-vdw': 4, 'scan': 5, 'tpss': 6, 'hse06': 7, 'pbe0': 8}
    functional_list = []
    for xc in functional_set:
        try: functional_list.append(Functional(functional_name=xc, slab_db=pd_solid_dat, adsorbate_db=None, mol_db=pd_molecule_dat, needed_struc_dict=need_structures, thermo_dynamic=True))
        except: pass

    functional_list = sorted(functional_list, key=lambda x: prefered_functional_order.get(x.name.lower()) if x.name.lower() in prefered_functional_order.keys() else 100)

    if verbose:
        for i, func in enumerate(functional_list):
            missing_structures(func, need_structures)

    O2_er = dict()
    N2_er = dict()
    NO_er = dict()
    CO_er = dict()

    for func in copy(functional_list):
        try:
            O2_er.update({func.name: func.calc_O2_err()})
            N2_er.update({func.name: func.calc_N2_err()})
            NO_er.update({func.name: func.calc_NO_err()})
            CO_er.update({func.name: func.calc_CO_err()})
            func.correct_energies()
        except: functional_list.remove(func)

    excel_file = xl.Workbook()
    work_sheet = excel_file.active
    work_sheet.title = 'energies'
#    deviation_sheet = excel_file.create_sheet('deviations')
    formation_sheet = excel_file.create_sheet('formation')
 #   formation_sheet_deviation = excel_file.create_sheet('formation_deviation')
    correction_sheet = excel_file.create_sheet('corrections simple')

    upper_border = Border(
        top=Side(border_style=BORDER_THIN, color='00000000'),
        )

    start_of_data = 3

    for i, func in enumerate(functional_list):
        for sheet in (work_sheet, formation_sheet): sheet.cell(1, i + start_of_data, func.name)
        correction_sheet.cell(1, i + start_of_data - 1, func.name)
        correction_sheet.cell(2, i + 2, O2_er[func.name])
        correction_sheet.cell(3, i + 2, N2_er[func.name])
        correction_sheet.cell(4, i + 2, NO_er[func.name])
        correction_sheet.cell(5, i + 2, CO_er[func.name])
    correction_sheet.cell(2, 1, 'O2 error')
    correction_sheet.cell(3, 1, 'N2 error')
    correction_sheet.cell(4, 1, 'NO error')
    correction_sheet.cell(5, 1, 'CO error')

    work_sheet.cell(1, len(functional_list) + start_of_data, 'exp ref')
    formation_sheet.cell(1, len(functional_list) + start_of_data, 'exp ref')

    next_row = 2

    for name, reactions in all_gaseous_reactions_named.items():
        for sheet in [work_sheet,]:
            sheet.cell(next_row, start_of_data - 2, name)
            for j in range(len(functional_list)*2+start_of_data+3): sheet.cell(next_row, j + 1).border = upper_border
        for i, reac in enumerate(reactions):
            for sheet in [work_sheet,]: sheet.cell(next_row, start_of_data - 1, reac.products[0].name)
            for j, func in enumerate(functional_list):
                if work_sheet.cell(1, j + start_of_data).value is None: work_sheet.cell(1, j + start_of_data, func.name)
                if work_sheet.cell(1, j + start_of_data + len(functional_list) + 2).value is None: work_sheet.cell(1, j + start_of_data + len(functional_list) + 2, func.name)
                try:
                    work_sheet.cell(next_row, j + start_of_data, func.calculate_reaction_enthalpy(reac))
                    work_sheet.cell(next_row, j + start_of_data + len(functional_list) + 2, func.calculate_reaction_enthalpy(reac) - reac.experimental_ref)

#                    deviation_sheet.cell(next_row, j + start_of_data, func.calculate_reaction_enthalpy(reac) - reac.experimental_ref)

                except: pass
            work_sheet.cell(next_row, len(functional_list) + start_of_data, reac.experimental_ref)
            next_row += 1

    #f'${xl.utils.cell.get_column_letter(start_of_data)}${start_of_data}:${xl.utils.cell.get_column_letter(j+start_of_data)}${next_row - 1}'
    for i, func in enumerate(functional_list):
        work_sheet.conditional_formatting.add((cond_zone := f'${xl.utils.cell.get_column_letter(start_of_data + len(functional_list) + 2 + i)}${2}:${xl.utils.cell.get_column_letter(start_of_data + len(functional_list) + 2 + i)}${next_row - 1}'),
                                               ColorScaleRule(start_type='formula',
                                                              start_value=f'-MAX(-MIN({cond_zone}),MAX({cond_zone}))',
                                                              start_color='AA0000',

                                                              mid_type='num',
                                                              mid_value=0,
                                                              mid_color='FFFFFF',

                                                              end_type='formula',
                                                              end_value=f'MAX(-MIN({cond_zone}),MAX({cond_zone}))',
                                                              end_color='AA0000'
                                                              ))

        cond_zone = cond_zone.replace('$', '')
        if work_sheet.cell(1, i + start_of_data + len(functional_list) * 2 + 4).value is None: work_sheet.cell(1, i + start_of_data + len(functional_list) * 2 + 4, func.name)
        if work_sheet.cell(2, start_of_data + len(functional_list) * 2 + 3).value is None: work_sheet.cell(2, start_of_data + len(functional_list) * 2 + 3, 'MAE')
        if work_sheet.cell(3, start_of_data + len(functional_list) * 2 + 3).value is None: work_sheet.cell(3, start_of_data + len(functional_list) * 2 + 3, 'MAX')
        work_sheet.cell(2, i + start_of_data + len(functional_list) * 2 + 4, f'=AVERAGE(ABS({cond_zone}))')
        work_sheet.cell(3, i + start_of_data + len(functional_list) * 2 + 4, f'=MAX(ABS({cond_zone}))')

    next_row = 2
    for name, reactions in all_formation_reactions_named.items():
        for sheet in [formation_sheet,]:
            sheet.cell(next_row, start_of_data - 2, name)
            for j in range(len(functional_list)*2+start_of_data+3): sheet.cell(next_row, j + 1).border = upper_border
        for i, reac in enumerate(reactions):
            for sheet in [formation_sheet,]: sheet.cell(next_row, start_of_data - 1, reac.products[0].name)
            for j, func in enumerate(functional_list):
                if formation_sheet.cell(1, j + start_of_data + len(functional_list) + 2).value is None: formation_sheet.cell(1, j + start_of_data + len(functional_list) + 2, func.name)
                try:
                    formation_sheet.cell(next_row, j + start_of_data, func.calculate_reaction_enthalpy(reac))
                    formation_sheet.cell(next_row, j + start_of_data + len(functional_list) + 2, func.calculate_reaction_enthalpy(reac) - reac.experimental_ref)

#                    formation_sheet_deviation.cell(next_row, j + start_of_data, func.calculate_reaction_enthalpy(reac) - reac.experimental_ref)

                except: pass
            formation_sheet.cell(next_row, len(functional_list) + start_of_data, reac.experimental_ref)
            next_row += 1

    for i, func in enumerate(functional_list):
        formation_sheet.conditional_formatting.add((cond_zone := f'${xl.utils.cell.get_column_letter(start_of_data + len(functional_list) + 2 + i)}${2}:${xl.utils.cell.get_column_letter(start_of_data + len(functional_list) + 2 + i)}${next_row - 1}'),
                                               ColorScaleRule(start_type='formula',
                                                              start_value=f'-MAX(-MIN({cond_zone}),MAX({cond_zone}))',
                                                              start_color='AA0000',

                                                              mid_type='num',
                                                              mid_value=0,
                                                              mid_color='FFFFFF',

                                                              end_type='formula',
                                                              end_value=f'MAX(-MIN({cond_zone}),MAX({cond_zone}))',
                                                              end_color='AA0000'
                                                              ))

        cond_zone = cond_zone.replace('$', '')
        if formation_sheet.cell(1, i + start_of_data + len(functional_list) * 2 + 4).value is None: formation_sheet.cell(1, i + start_of_data + len(functional_list) * 2 + 4, func.name)
        if formation_sheet.cell(2, start_of_data + len(functional_list) * 2 + 3).value is None: formation_sheet.cell(2, start_of_data + len(functional_list) * 2 + 3, 'MAE')
        if formation_sheet.cell(3, start_of_data + len(functional_list) * 2 + 3).value is None: formation_sheet.cell(3, start_of_data + len(functional_list) * 2 + 3, 'MAX')
        formation_sheet.cell(2, i + start_of_data + len(functional_list) * 2 + 4, f'=AVERAGE(ABS({cond_zone}))')
        formation_sheet.cell(3, i + start_of_data + len(functional_list) * 2 + 4, f'=MAX(ABS({cond_zone}))')

    reac_group_names = {'Simple alkanes': 'methyl_carbons', 'Branched alkanes (iso)': 'iso_carbons', 'Branched alkanes (neo)': 'neo_carbons', 'Amines': 'amines', 'Nitros': 'nitro', 'Nitrates': 'nitrate', 'Nitrites': 'nitrite', 'Hydroxylamines': 'hydroxylamine', 'Aromatics': 'phenyl', 'Anilines': 'aniline', 'Hydrazines': 'hydrazine', 'Amides': 'amide', 'Nitriles': 'nitrile'}

    correction_sheet_gas_simple = excel_file.create_sheet('corrected gas simple')
    correction_sheet_form_simple = excel_file.create_sheet('corrected formation simple')
    start_of_simple_corr = 8

    for i, func in enumerate(functional_list):
        try:
            corrections = simple_decomposition(func, molecule_functional_dict)
            for j, (group_name, corr_key) in enumerate(reac_group_names.items()):
                if correction_sheet.cell(start_of_simple_corr + j, 1).value is None: correction_sheet.cell(start_of_simple_corr + j, 1, group_name)
                correction_sheet.cell(start_of_simple_corr + j, 2 + i, corrections[corr_key])
            for sheet in [correction_sheet_gas_simple, correction_sheet_form_simple]: sheet.cell(1, 3 + i, func.name)
            for sheet, reac_group_dict in ((correction_sheet_gas_simple, all_gaseous_reactions_named), (correction_sheet_form_simple, all_formation_reactions_named)):
                next_row = 2
                for reac_group_name, reac_list in reac_group_dict.items():
                    if sheet.cell(next_row, 1).value is None: sheet.cell(next_row, 1, reac_group_name)
                    for j in range(start_of_simple_corr + 2 * len(functional_list)): sheet.cell(next_row, j + 1).border = upper_border
                    for reac in reac_list:
                        if sheet.cell(next_row, 2).value is None: sheet.cell(next_row, 2, reac.products[0].name)
                        sheet.cell(next_row, 3 + i, func.calculate_reaction_enthalpy(reac)
                                                 + sum(sum(corrections[f_group] * f_amount for f_group, f_amount in molecule_functional_dict[comp.name].items() if f_group in corrections.keys()) * comp.amount for comp in reac.reactants if comp.name in molecule_functional_dict.keys())
                                                 - sum(sum(corrections[f_group] * f_amount for f_group, f_amount in molecule_functional_dict[comp.name].items() if f_group in corrections.keys()) * comp.amount for comp in reac.products if comp.name in molecule_functional_dict.keys()))

                        if sheet.cell(1, 3 + i + len(functional_list) + 2).value is None: sheet.cell(1, 3 + i + len(functional_list) + 2, func.name)
                        sheet.cell(next_row, 3 + i + len(functional_list) + 2, func.calculate_reaction_enthalpy(reac)
                                                 + sum(sum(corrections[f_group] * f_amount for f_group, f_amount in molecule_functional_dict[comp.name].items() if f_group in corrections.keys()) * comp.amount for comp in reac.reactants if comp.name in molecule_functional_dict.keys())
                                                 - sum(sum(corrections[f_group] * f_amount for f_group, f_amount in molecule_functional_dict[comp.name].items() if f_group in corrections.keys()) * comp.amount for comp in reac.products if comp.name in molecule_functional_dict.keys())
                                                 - reac.experimental_ref)
                        if i == 0:
                            sheet.cell(1, 3 + len(functional_list), 'exp ref')
                            sheet.cell(next_row, 3 + len(functional_list), reac.experimental_ref)
                        next_row += 1
        except: pass

    correction_sheet_linalg = excel_file.create_sheet('corrections linalg')
    correction_sheet_gas_linalg = excel_file.create_sheet('corrected gas linalg')
    correction_sheet_form_linalg = excel_file.create_sheet('corrected formation linalg')
    correction_sheet_linalg.cell(2, 1, 'residual')
    start_of_linalg_data = 5

    for i, func in enumerate(functional_list):
        try:
            for sheet in [correction_sheet_linalg,]: sheet.cell(1, 2 + i, func.name)
            for sheet in [correction_sheet_gas_linalg, correction_sheet_form_linalg]: sheet.cell(1, 3 + i, func.name)
            corrections, residual = lstsq_decomposition(func, all_formation_reactions, molecule_functional_dict)
            correction_sheet_linalg.cell(2, 2 + i, residual)
            for j, (group_name, corr_key) in enumerate(reac_group_names.items()):
                if correction_sheet_linalg.cell(start_of_linalg_data + j, 1).value is None: correction_sheet_linalg.cell(start_of_linalg_data + j, 1, group_name)
                correction_sheet_linalg.cell(start_of_linalg_data + j, 2 + i,  corrections[corr_key])
            for sheet, reac_group_dict in ((correction_sheet_gas_linalg, all_gaseous_reactions_named), (correction_sheet_form_linalg, all_formation_reactions_named)):
                next_row = 2
                for reac_group_name, reac_list in reac_group_dict.items():
                    if sheet.cell(next_row, 1).value is None: sheet.cell(next_row, 1, reac_group_name)
                    for j in range(5 + 2 * len(functional_list)): sheet.cell(next_row, j + 1).border = upper_border
                    for reac in reac_list:
                        if sheet.cell(next_row, 2).value is None: sheet.cell(next_row, 2, reac.products[0].name)
                        sheet.cell(next_row, 3 + i, func.calculate_reaction_enthalpy(reac)
                                                 + sum(sum(corrections[f_group] * f_amount for f_group, f_amount in molecule_functional_dict[comp.name].items() if f_group in corrections.keys()) * comp.amount for comp in reac.reactants if comp.name in molecule_functional_dict.keys())
                                                 - sum(sum(corrections[f_group] * f_amount for f_group, f_amount in molecule_functional_dict[comp.name].items() if f_group in corrections.keys()) * comp.amount for comp in reac.products if comp.name in molecule_functional_dict.keys()))

                        if sheet.cell(1, 3 + i + len(functional_list) + 2).value is None: sheet.cell(1, 3 + i + len(functional_list) + 2, func.name)
                        sheet.cell(next_row, 3 + i + len(functional_list) + 2, func.calculate_reaction_enthalpy(reac)
                                                 + sum(sum(corrections[f_group] * f_amount for f_group, f_amount in molecule_functional_dict[comp.name].items() if f_group in corrections.keys()) * comp.amount for comp in reac.reactants if comp.name in molecule_functional_dict.keys())
                                                 - sum(sum(corrections[f_group] * f_amount for f_group, f_amount in molecule_functional_dict[comp.name].items() if f_group in corrections.keys()) * comp.amount for comp in reac.products if comp.name in molecule_functional_dict.keys())
                                                 - reac.experimental_ref)
                        if i == 0:
                            sheet.cell(1, 3 + len(functional_list), 'exp ref')
                            sheet.cell(next_row, 3 + len(functional_list), reac.experimental_ref)
                        next_row += 1
        except: pass

    for sheet, reac_group in ((correction_sheet_gas_linalg, all_gaseous_reactions), (correction_sheet_gas_simple, all_gaseous_reactions), (correction_sheet_form_linalg, all_formation_reactions), (correction_sheet_form_simple, all_formation_reactions)):

        for i, func in enumerate(functional_list):
            sheet.conditional_formatting.add((cond_zone := f'${xl.utils.cell.get_column_letter(2 + len(functional_list) + 3 + i)}${2}:${xl.utils.cell.get_column_letter(2 + len(functional_list) + 3 + i)}${len(reac_group) + 1}'),
                                       ColorScaleRule(start_type='formula',
                                                      start_value=f'-MAX(-MIN({cond_zone}),MAX({cond_zone}))',
                                                      start_color='AA0000',

                                                      mid_type='num',
                                                      mid_value=0,
                                                      mid_color='FFFFFF',

                                                      end_type='formula',
                                                      end_value=f'MAX(-MIN({cond_zone}),MAX({cond_zone}))',
                                                      end_color='AA0000'
                                                      ))

            cond_zone = cond_zone.replace('$','')
            if sheet.cell(1, i + start_of_data + len(functional_list) * 2 + 4).value is None: sheet.cell(1, i + start_of_data + len(functional_list) * 2 + 4, func.name)
            if sheet.cell(2,start_of_data + len(functional_list) * 2 + 3).value is None: sheet.cell(2,start_of_data + len(functional_list) * 2 + 3, 'MAE')
            if sheet.cell(3, start_of_data + len(functional_list) * 2 + 3).value is None: sheet.cell(3, start_of_data + len(functional_list) * 2 + 3, 'MAX')
            sheet.cell(2, i + start_of_data + len(functional_list) * 2 + 4, f'=AVERAGE(ABS({cond_zone}))')
            sheet.cell(3, i + start_of_data + len(functional_list) * 2 + 4, f'=MAX(ABS({cond_zone}))')

    excel_file.save('reaction_results.xlsx')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('molecule_db_directory', help='Path to the database')
    parser.add_argument('solid_db_directory', help='Path to the database')
    parser.add_argument('--verbose', action='store_true')
    args = parser.parse_args()

    main(args.molecule_db_directory, args.solid_db_directory, args.verbose)
